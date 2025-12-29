"""Dropout counts extractor for the foundation pipeline."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Required, TypedDict

import polars as pl
from openpyxl import load_workbook

from ..common import console
from ..plugin import BaseExtractor, ExtractionContext, ExtractionResult


class DropoutSheetConfig(TypedDict, total=False):
    """TypedDict describing how to read each dropout workbook."""

    year: Required[str]
    sheet_name: Required[str]
    cols: Required[str]
    schema: Required[str]
    school_id_col: Required[str]
    header: int
    header_rows: int
    header_start: int
    data_start: int


class _DropoutSheetConfigRequired(TypedDict):
    """Required keys for DropoutSheetConfig."""

    sheet_name: str
    school_id_col: str
    schema: str
    year: str


DROP_OUT_CONFIGS: list[DropoutSheetConfig] = [
    {
        "year": "2022-2023",
        "sheet_name": "DROPOUT DB",
        "cols": "E,K:Z,AD:AM,AQ:AX,BB:BI",
        "header": 6,
        "school_id_col": "School ID",
        "schema": "2022",
    },
    {
        "year": "2023-2024",
        "sheet_name": "DB DROPOUT",
        "cols": "E,K:Z,AD:AM,AQ:AT",
        "header_rows": 2,
        "header_start": 3,
        "data_start": 7,
        "school_id_col": "school_id",
        "schema": "2023",
    },
    {
        "year": "2024-2025",
        "sheet_name": "DATABASE",
        "cols": "C,K:Z,AD:AM,AQ:AT",
        "header": 8,
        "school_id_col": "BEIS School ID",
        "schema": "2024",
    },
]


_DROP_OUT_COLUMNS = {
    "school_year": pl.Utf8,
    "school_id": pl.Utf8,
    "grade": pl.Utf8,
    "strand": pl.Utf8,
    "sex": pl.Utf8,
    "num_dropouts": pl.Int64,
    "source_file": pl.Utf8,
    "source_row": pl.Int64,
    "ingested_at": pl.Datetime,
}


def _excel_column_index(value: str) -> int:
    """Convert an Excel column letter to a zero-based index."""

    value = value.strip().upper()
    index = 0
    for char in value:
        if not char.isalpha():
            raise ValueError(f"Invalid Excel column reference: {value}")
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _parse_usecols(usecols: str | None) -> list[int] | None:
    if not usecols:
        return None

    indexes: set[int] = set()
    for part in usecols.split(","):
        segment = part.strip()
        if not segment:
            continue
        if ":" in segment:
            start, end = [token.strip() for token in segment.split(":", 1)]
            indexes.update(
                range(_excel_column_index(start), _excel_column_index(end) + 1)
            )
        else:
            indexes.add(_excel_column_index(segment))
    return sorted(indexes)


def _normalize_cell_value(value: object) -> object | None:
    if value in ("#N/A", None, ""):
        return None
    return value


def _normalize_header(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _forward_fill_row(row: tuple[object | None, ...]) -> list[str | None]:
    filled: list[str | None] = []
    last: str | None = None
    for cell in row:
        normalized = _normalize_header(cell)
        if normalized:
            last = normalized
        filled.append(last)
    return filled


def _build_headers(
    header_row: tuple[object | None, ...], indexes: list[int]
) -> list[str]:
    headers: list[str] = []
    for idx in indexes:
        if idx < len(header_row):
            value = _normalize_header(header_row[idx])
        else:
            value = None
        headers.append(value or f"column_{idx}")
    return _ensure_unique_column_names(headers)


def _ensure_unique_column_names(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_headers: list[str] = []
    for label in headers:
        count = seen.get(label, 0)
        if count:
            unique_headers.append(f"{label}_{count}")
            seen[label] = count + 1
        else:
            unique_headers.append(label)
            seen[label] = 1
    return unique_headers


def _build_headers_from_rows(
    header_rows: list[tuple[object | None, ...]], indexes: list[int]
) -> list[str]:
    filled_rows = [_forward_fill_row(row) for row in header_rows]
    headers: list[str] = []
    for idx in indexes:
        tokens: list[str] = []
        for row in filled_rows:
            if idx < len(row) and row[idx]:
                tokens.append(row[idx])  # type: ignore[arg-type]
        combined = " ".join(tokens)
        text = re.sub(r"\s+", " ", combined).strip()
        headers.append(text or f"column_{idx}")
    return _ensure_unique_column_names(headers)


def _read_simple_sheet(path: Path, cfg: DropoutSheetConfig) -> pl.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = cfg.get("sheet_name")
    if not sheet_name:
        raise ValueError("sheet_name is required in config")
    try:
        ws = wb[sheet_name]
    except KeyError as exc:
        raise ValueError(f"Missing sheet {sheet_name}") from exc

    rows = list(ws.iter_rows(values_only=True))
    header_index = cfg.get("header")
    if header_index is None or header_index >= len(rows):
        raise ValueError(f"Header index {header_index} out of range for {path.name}")

    header_row = rows[header_index]
    selected_indexes = _parse_usecols(cfg.get("cols")) or list(range(len(header_row)))
    headers = _build_headers(header_row, selected_indexes)
    school_id_col = cfg.get("school_id_col")
    if school_id_col:
        headers[0] = school_id_col

    data_start = header_index + 1
    data_values: list[list[object | None]] = []
    for offset, row in enumerate(rows[data_start:], start=1):
        row_cells = []
        for idx in selected_indexes:
            value = row[idx] if idx < len(row) else None
            row_cells.append(_normalize_cell_value(value))
        row_cells.append(data_start + offset)
        data_values.append(row_cells)

    columns = headers + ["__source_row"]
    return pl.DataFrame(data=data_values, schema=columns, orient="row")


def _read_merged_headers_sheet(path: Path, cfg: DropoutSheetConfig) -> pl.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = cfg.get("sheet_name")
    if not sheet_name:
        raise ValueError("sheet_name is required in config")
    try:
        ws = wb[sheet_name]
    except KeyError as exc:
        raise ValueError(f"Missing sheet {sheet_name}") from exc

    rows = list(ws.iter_rows(values_only=True))
    header_start = cfg.get("header_start")
    header_rows = cfg.get("header_rows")
    if header_start is None or header_rows is None:
        raise ValueError("Merged-sheet config missing header info")

    header_slice = rows[header_start : header_start + header_rows]
    if not header_slice:
        raise ValueError(f"Not enough header rows for {path.name}")
    first_header = header_slice[0]
    selected_indexes = _parse_usecols(cfg.get("cols")) or list(range(len(first_header)))
    normalized_header_slice: list[tuple[object | None, ...]] = [
        tuple(_normalize_cell_value(cell) for cell in row) for row in header_slice
    ]
    headers = _build_headers_from_rows(normalized_header_slice, selected_indexes)
    school_id_col = cfg.get("school_id_col")
    if school_id_col:
        headers[0] = school_id_col

    data_start = cfg.get("data_start")
    if data_start is None:
        raise ValueError("Merged-sheet config missing data_start")

    data_values: list[list[object | None]] = []
    for offset, row in enumerate(rows[data_start:], start=1):
        row_cells = []
        for idx in selected_indexes:
            value = row[idx] if idx < len(row) else None
            row_cells.append(_normalize_cell_value(value))
        row_cells.append(data_start + offset)
        data_values.append(row_cells)

    columns = headers + ["__source_row"]
    return pl.DataFrame(data=data_values, schema=columns, orient="row")


def _read_dropout_df(path: Path, cfg: DropoutSheetConfig) -> pl.DataFrame:
    if "header" in cfg:
        return _read_simple_sheet(path, cfg)
    return _read_merged_headers_sheet(path, cfg)


def _melt_dropout(
    df: pl.DataFrame,
    *,
    school_id_col: str,
    school_year: str,
    source_file: str,
    ingested_at: datetime,
) -> pl.DataFrame:
    id_vars = [school_id_col]
    if "__source_row" in df.columns:
        id_vars.append("__source_row")

    melted = df.unpivot(
        index=id_vars,
        value_name="num_dropouts",
        variable_name="raw_col",
    ).with_columns(
        pl.col("num_dropouts").cast(pl.Float64, strict=False).alias("num_dropouts")
    )

    melted = (
        melted.filter(pl.col("num_dropouts").is_not_null())
        .filter(pl.col("num_dropouts") > 0)
        .rename({school_id_col: "school_id", "__source_row": "source_row"})
        .with_columns(
            pl.lit(school_year).alias("school_year"),
            pl.lit(source_file).alias("source_file"),
            pl.lit(ingested_at).alias("ingested_at"),
        )
    )

    if "source_row" in melted.columns:
        melted = melted.with_columns(pl.col("source_row").cast(pl.Int64, strict=False))
    return melted


def _normalize_school_id(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0"):
        text = text[: -len(".0")]
    return text


def _parse_numeric(value: object | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value != value:  # NaN
            return None
        return int(value)
    cleaned = str(value).strip()
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", "")
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def _parse_dropout_records(
    df: pl.DataFrame,
    parser: Callable[[str], tuple[str | None, str | None, str | None]],
) -> tuple[list[dict[str, object | None]], list[dict[str, object | None]]]:
    records: list[dict[str, object | None]] = []
    invalid: list[dict[str, object | None]] = []
    for row in df.to_dicts():
        grade, strand, sex = parser(row["raw_col"])
        if not grade or not sex:
            invalid.append(row)
            continue
        school_id = _normalize_school_id(row["school_id"])
        if not school_id:
            invalid.append(row)
            continue
        num_dropouts = _parse_numeric(row["num_dropouts"])
        if num_dropouts is None:
            invalid.append(row)
            continue

        records.append(
            {
                "school_year": row["school_year"],
                "school_id": school_id,
                "grade": grade,
                "strand": strand,
                "sex": sex,
                "num_dropouts": num_dropouts,
                "source_file": row["source_file"],
                "source_row": row.get("source_row"),
                "ingested_at": row["ingested_at"],
            }
        )
    return records, invalid


def _parser_for_schema(
    schema: str,
) -> Callable[[str], tuple[str | None, str | None, str | None]]:
    parsers = {
        "2022": _parse_dropout_2022,
        "2023": _parse_dropout_2023,
        "2024": _parse_dropout_2024,
    }
    try:
        return parsers[schema]
    except KeyError as exc:
        raise ValueError(f"No parser for dropout schema {schema}") from exc


def consolidate_dropouts(folder: Path) -> tuple[pl.DataFrame, dict[str, object]]:
    ingested_at = datetime.now(timezone.utc)
    records: list[dict[str, object | None]] = []
    invalid_rows: list[dict[str, object | None]] = []
    melted_rows = 0
    processed_files = 0

    for cfg in DROP_OUT_CONFIGS:
        year = cfg.get("year")
        if not year:
            raise ValueError("Missing 'year' in dropout config")
        path = folder / f"{year}-dropouts.xlsx"
        if not path.exists():
            raise FileNotFoundError(f"Dropout file missing: {path}")
        df_raw = _read_dropout_df(path, cfg)
        parser = _parser_for_schema(cfg["schema"])
        melted = _melt_dropout(
            df_raw,
            school_id_col=cfg["school_id_col"],
            school_year=year,
            source_file=path.name,
            ingested_at=ingested_at,
        )
        melted_rows += melted.height
        processed_files += 1

        parsed, invalid = _parse_dropout_records(melted, parser)
        records.extend(parsed)
        invalid_rows.extend(invalid)

    if records:
        df = pl.DataFrame(records, schema=_DROP_OUT_COLUMNS)
    else:
        df = pl.DataFrame(schema=_DROP_OUT_COLUMNS)

    before = df.height
    df = df.unique(
        subset=["school_year", "school_id", "grade", "strand", "sex"], keep="last"
    )
    duplicates_removed = before - df.height

    metrics: dict[str, object] = {
        "dropouts_files": processed_files,
        "dropouts_rows_melted": melted_rows,
        "dropouts_valid_rows": df.height,
        "dropouts_invalid_rows": len(invalid_rows),
        "dropouts_duplicates_removed": duplicates_removed,
    }

    if invalid_rows:
        sample = invalid_rows[:3]
        console.log(
            f"[yellow]Dropped {len(invalid_rows)} invalid dropout rows; sample:[/yellow] {sample}"
        )

    console.log(
        f"[cyan]Dropout ingest[/cyan] files={processed_files} "
        f"melted={melted_rows} valid={metrics['dropouts_valid_rows']} "
        f"invalid={metrics['dropouts_invalid_rows']} duplicates={duplicates_removed}"
    )

    return df, metrics


def _parse_dropout_2022(col: str) -> tuple[str | None, str | None, str | None]:
    m = re.search(r"(Male|Female)", col)
    if not m:
        return None, None, None
    sex = m.group(1).lower()[0]

    base = re.sub(r"\s+(Male|Female)\s+Dropout$", "", col).strip()

    if base == "K":
        return "kinder", None, sex
    if base.startswith("NG Elem"):
        return "esng", None, sex
    if base.startswith("NG Second"):
        return "jhsng", None, sex

    match = re.match(r"(G1[12])\s+(.+)", base)
    if match:
        grade = match.group(1).lower()
        raw_strand = match.group(2).strip().lower()
        strand_match = re.match(r"acad\s*-?\s*(.+)", raw_strand)
        if strand_match:
            strand = strand_match.group(1)
        else:
            strand = raw_strand
        return grade, strand, sex

    match = re.match(r"G([1-9]|10)", base)
    if match:
        return f"g{match.group(1)}", None, sex

    return None, None, None


def _parse_dropout_2023(col: str) -> tuple[str | None, str | None, str | None]:
    lower = col.lower()
    if lower.endswith("male"):
        sex = "m"
    elif lower.endswith("female"):
        sex = "f"
    else:
        return None, None, None

    if lower.startswith("kindergarten"):
        return "kinder", None, sex

    if lower.startswith("non-graded"):
        return "esng", None, sex

    match = re.match(r"grade\s+(1[0-2]|[1-9])", lower)
    if match:
        return f"g{match.group(1)}", None, sex

    return None, None, None


def _parse_dropout_2024(col: str) -> tuple[str | None, str | None, str | None]:
    match = re.match(r"dropout_([^_]+)_(male|female)", col.lower())
    if not match:
        return None, None, None

    raw_grade, raw_sex = match.groups()
    sex = "m" if raw_sex == "male" else "f"

    if raw_grade == "k":
        return "kinder", None, sex
    if raw_grade == "ng":
        return "esng", None, sex
    if raw_grade.startswith("g"):
        return raw_grade, None, sex

    return None, None, None


class DropoutsExtractor(BaseExtractor):
    """Consume standardized dropout workbooks and expose a `dropouts` facts table."""

    name = "dropouts"
    depends_on = ["school_year_meta", "enrollment", "school_levels"]
    outputs = ["dropouts"]

    def extract(
        self,
        context: ExtractionContext,
        dependencies: dict[str, pl.DataFrame],
    ) -> ExtractionResult:
        del dependencies

        dropout_dir = context.paths.dropout_dir
        if not dropout_dir.exists():
            raise FileNotFoundError(f"Dropout directory {dropout_dir} does not exist")

        df, metrics = consolidate_dropouts(dropout_dir)
        if df.height == 0:
            raise ValueError("Dropout extraction produced no rows")

        return ExtractionResult(tables={"dropouts": df}, metrics=metrics)
