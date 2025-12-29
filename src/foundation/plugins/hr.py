"""Teacher headcount extractor that joins the HR folder into the pipeline."""

from __future__ import annotations

import re
from pathlib import Path
from typing import TypedDict

import polars as pl
from openpyxl import load_workbook

from ..plugin import BaseExtractor, ExtractionContext, ExtractionResult


class SheetConfig(TypedDict):
    """TypedDict for workbook sheet metadata."""

    sheet_name: str
    usecols: str | None
    header: int | None


YEAR_SHEET_CONFIGS: dict[str, dict[str, SheetConfig]] = {
    "2022-2023": {
        "es": {"sheet_name": "ES DB", "usecols": "D:E,S:AD", "header": 5},
        "jhs": {"sheet_name": "JHS DB", "usecols": "D:E,S:AH", "header": 5},
        "shs": {"sheet_name": "SHS DB", "usecols": "D:E,T:AA", "header": 5},
    },
    "2023-2024": {
        "es": {"sheet_name": "ES DB", "usecols": "D,O:Z", "header": 6},
        "jhs": {"sheet_name": "JHS DB", "usecols": "D,O:AD", "header": 6},
        "shs": {"sheet_name": "SHS", "usecols": "D,O:V", "header": 7},
    },
    "2024-2025": {
        "es": {"sheet_name": "ES", "usecols": "E:F,U:AF", "header": 7},
        "jhs": {"sheet_name": "JHS", "usecols": "E:F,U:AJ", "header": 7},
        "shs": {"sheet_name": "SHS", "usecols": "E:F,U:AB", "header": 7},
    },
}


def _excel_column_index(value: str) -> int:
    """Convert Excel column letters to a zero-based index."""

    value = value.strip().upper()
    index = 0
    for char in value:
        if not char.isalpha():
            raise ValueError(f"Invalid Excel column reference: {value}")
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _parse_usecols(usecols: str | None) -> list[int] | None:
    if not usecols:
        return None

    indexes: set[int] = set()
    for part in usecols.split(","):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            start, end = [p.strip() for p in part.split(":")]
            indexes.update(
                range(_excel_column_index(start), _excel_column_index(end) + 1)
            )
        else:
            indexes.add(_excel_column_index(part))
    return sorted(indexes)


def _normalize_school_id(df: pl.DataFrame) -> pl.DataFrame:
    """Create a canonical ``school_id`` column based on LIS/BEIS IDs."""

    column_map = {col: str(col).strip().lower() for col in df.columns}
    df = df.rename(column_map)

    lis_col = "lis school id"
    beis_col = "beis school id"
    cols = set(df.columns)

    if lis_col in cols and beis_col in cols:
        expr = pl.when(pl.col(lis_col).is_not_null())
        expr = expr.then(pl.col(lis_col)).otherwise(pl.col(beis_col))
        df = df.with_columns(school_id=expr)
        df = df.drop([lis_col, beis_col])
    elif lis_col in cols:
        df = df.with_columns(school_id=pl.col(lis_col))
        df = df.drop(lis_col)
    elif beis_col in cols:
        df = df.with_columns(school_id=pl.col(beis_col))
        df = df.drop(beis_col)
    else:
        raise ValueError(f"No {lis_col=}/{beis_col=} found. See {list(df.columns)=}")

    cleaned_id = pl.col("school_id").cast(pl.Utf8).str.replace_all(r"\D", "")
    df = df.with_columns(
        school_id=pl.when(cleaned_id.str.len_chars().eq(pl.lit(0)))
        .then(pl.lit("0"))
        .otherwise(cleaned_id)
    )
    df = df.with_columns(school_id=pl.col("school_id").cast(pl.Int64))

    return df


def _melt_teacher_counts(df: pl.DataFrame, level: str) -> pl.DataFrame:
    """Unpivot the positional columns so each row represents a headcount."""

    df_long = df.unpivot(
        index="school_id",
        variable_name="position",
        value_name="num",
    )
    df_long = df_long.with_columns(level=pl.lit(level))
    df_long = df_long.with_columns(
        position=pl.col("position").str.replace_all(
            r"^sped teacher\s+([iv]+)$",
            r"sped/sned teacher \1",
            literal=False,
        )
    )
    df_long = df_long.with_columns(num=pl.col("num").cast(pl.Int64, strict=False))
    df_long = df_long.filter(pl.col("num").is_not_null() & (pl.col("num") != 0))
    return df_long


def _read_sheet(workbook: Path, cfg: SheetConfig, level: str) -> pl.DataFrame:
    """Use openpyxl to extract the exact header row and columns."""

    start_row = cfg["header"] or 0
    column_indexes = _parse_usecols(cfg["usecols"])

    wb = load_workbook(workbook, data_only=True, read_only=True)
    try:
        ws = wb[cfg["sheet_name"]]
    except KeyError as exc:
        raise ValueError(f"Sheet {cfg['sheet_name']} is missing") from exc

    rows = list(ws.iter_rows(values_only=True))
    if start_row >= len(rows):
        raise ValueError(
            f"Header index {start_row} is beyond sheet height {len(rows)} for {cfg['sheet_name']}"
        )

    header_row = rows[start_row]
    selected_indexes = column_indexes or list(range(len(header_row)))

    header = []
    for idx in selected_indexes:
        value = header_row[idx] if idx < len(header_row) else None
        header.append(
            str(value).strip() if value not in (None, "") else f"column_{idx}"
        )

    data_rows = rows[start_row + 1 :]
    values: list[list[object | None]] = []
    for row in data_rows:
        values.append(
            [
                None if idx >= len(row) else (None if row[idx] == "#N/A" else row[idx])
                for idx in selected_indexes
            ]
        )

    df = pl.DataFrame(values, schema=header, orient="row")
    df = df.with_columns([pl.col(col).cast(pl.Utf8) for col in header])
    df = _normalize_school_id(df)
    return _melt_teacher_counts(df, level)


def _load_teacher_sheets(
    workbook: Path, sheet_configs: dict[str, SheetConfig]
) -> pl.DataFrame:
    """Read the configured sheets and combine them into a single DataFrame."""

    frames: list[pl.DataFrame] = []
    for level, cfg in sheet_configs.items():
        frames.append(_read_sheet(workbook, cfg, level))

    if not frames:
        return pl.DataFrame(schema={"school_year": pl.Utf8})

    return pl.concat(frames, how="vertical")


def _read_teacher_file(file: Path) -> pl.DataFrame:
    """Load a single Excel workbook and tag every row with its school year."""

    match = re.search(r"\d{4}-\d{4}", file.name)
    if not match:
        raise ValueError(f"Cannot detect school year on {file.name}")

    year_range = match.group(0)
    sheet_configs = YEAR_SHEET_CONFIGS.get(year_range)
    if not sheet_configs:
        raise ValueError(f"No loader defined for school year: {year_range}")

    df = _load_teacher_sheets(file, sheet_configs)
    if df.is_empty():
        return df.with_columns(school_year=pl.lit(year_range))

    return df.with_columns(school_year=pl.lit(year_range)).select(
        ["school_year", "school_id", "level", "position", "num"]
    )


class TeachersExtractor(BaseExtractor):
    """Load teacher headcount Excel files and expose them to the pipeline."""

    name = "teachers"
    depends_on = ["school_year_meta", "enrollment", "school_levels"]
    outputs = ["teachers"]

    def extract(
        self,
        context: ExtractionContext,
        dependencies: dict[str, pl.DataFrame],
    ) -> ExtractionResult:
        hr_dir = context.paths.hr_dir
        if not hr_dir.exists():
            raise FileNotFoundError(f"HR directory {hr_dir} does not exist")

        files = sorted(hr_dir.glob("*.xlsx"))
        if not files:
            raise FileNotFoundError(f"No teacher Excel files found in {hr_dir}")

        frames = [_read_teacher_file(file) for file in files]
        combined = pl.concat(frames, how="vertical")
        if combined.is_empty():
            raise ValueError("Teacher load produced no rows")

        return ExtractionResult(tables={"teachers": combined})
